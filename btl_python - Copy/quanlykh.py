from tkinter import *
from tkinter import ttk, messagebox
from PIL import ImageTk, Image
from tkcalendar import DateEntry
from datetime import datetime
import re
import database
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def show_KhachHang(frame_noidung):
    for widget in frame_noidung.winfo_children():
        widget.destroy()
    def display_data():
        try:
            rows = database.Hien_Thi_KH()
            update_table(rows)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi hiển thị dữ liệu: {str(e)}")

    def update_table(rows):
        table_kh.delete(*table_kh.get_children())
        for row in rows:
            table_kh.insert("", END, values=row)
    def clear_entries():
        input_ten_kh.delete(0, END)
        input_ngay_kh.set_date(datetime.now())  # Reset DateEntry to current date
        input_sex_kh.set('')
        input_sdt_kh.delete(0, END)
        input_cccd_kh.delete(0, END)
        input_email_kh.delete(0, END)
        input_diachi_kh.delete(0, END)

    def ChonDuLieuBang(event):
        selected_row = table_kh.focus()
        data = table_kh.item(selected_row, 'values')
        ma_chon = data[0]
        clear_entries()
        input_ten_kh.insert(0, data[1])
        input_ngay_kh.set_date(datetime.strptime(data[2], "%Y-%m-%d"))  # Set DateEntry value
        input_sex_kh.set(data[3])
        input_sdt_kh.insert(0, data[4])
        input_cccd_kh.insert(0, data[5])
        input_email_kh.insert(0, data[6])
        input_diachi_kh.insert(0, data[7])



    def sua():
        try:
            selected_row = table_kh.focus()
            data = table_kh.item(selected_row, 'values')
            ma_chon = data[0]
            ngay_sinh = input_ngay_kh.get_date().strftime("%Y-%m-%d")  # Convert DateEntry value to string
            database.sua_kh(ma_chon, input_ten_kh.get(), ngay_sinh, input_sex_kh.get(), int(input_sdt_kh.get()),
                            int(input_cccd_kh.get()), input_email_kh.get(), input_diachi_kh.get())
            messagebox.showinfo("Thông báo", "Cập nhật khách hàng thành công")
            clear_entries()
            display_data()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi cập nhật khách hàng: {str(e)}")

    def xoa():
        try:
            selected_row = table_kh.focus()
            data = table_kh.item(selected_row, 'values')
            ma_chon = data[0]
            database.xoa_kh(ma_chon)
            messagebox.showinfo("Thông báo", "Xóa khách hàng thành công")
            clear_entries()
            display_data()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xóa khách hàng: {str(e)}")

    def them(input_ten_kh, input_ngay_kh, input_sex_kh, input_sdt_kh, input_cccd_kh, input_email_kh, input_diachi_kh,
             addThem):
        # Lấy giá trị từ các trường nhập liệu
        ten_kh = input_ten_kh.get().strip()
        ngay_sinh = input_ngay_kh.get_date().strftime("%Y-%m-%d")  # Chuyển giá trị DateEntry thành chuỗi
        gioi_tinh = input_sex_kh.get().strip()
        sdt = input_sdt_kh.get().strip()
        cccd = input_cccd_kh.get().strip()
        email = input_email_kh.get().strip()
        dia_chi = input_diachi_kh.get().strip()

        # Kiểm tra điều kiện nhập liệu
        if ten_kh == "" or ngay_sinh == "" or gioi_tinh == "" or sdt == "" or cccd == ""or email == "" or dia_chi == "":
            messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ thông tin.")
            return
        if gioi_tinh not in ["Nam", "Nữ", "Khác"]:
            messagebox.showerror("Lỗi", "Giới tính phải chọn từ danh sách: Nam, Nữ, Khác.")
            return
        try:
            dob = datetime.strptime(ngay_sinh, "%Y-%m-%d")
            age = (datetime.now() - dob).days // 365
            if age < 18:
                messagebox.showerror("Lỗi", "Khách hàng phải từ 18 tuổi trở lên.")
                return
        except ValueError:
            messagebox.showerror("Lỗi", "Định dạng ngày sinh không hợp lệ. Định dạng đúng là YYYY-MM-DD.")
            return

        # Kiểm tra email: phải có đuôi @gmail.com
        if not email.endswith("@gmail.com"):
            messagebox.showerror("Lỗi", "Email phải có đuôi @gmail.com.")
            return
        # Kiểm tra CCCD: phải đủ 12 số và không được chứa chữ
        if len(cccd) != 12 or not cccd.isdigit():
            messagebox.showerror("Lỗi", "CCCD phải là số có 12 chữ số.")
            return
        # Kiểm tra nếu CCCD đã tồn tại trong cơ sở dữ liệu
        if database.check_cccd_exists(cccd):
            messagebox.showerror("Lỗi", "CCCD đã tồn tại.")
            return

        # Kiểm tra số điện thoại: phải là số Việt Nam
        if not sdt.isdigit() or len(sdt) not in [10, 11] or not sdt.startswith(('03', '05', '07', '08', '09')):
            messagebox.showerror("Lỗi", "Số điện thoại không hợp lệ. Phải là số Việt Nam và có 10 hoặc 11 chữ số.")
            return
        try:
            # Thêm khách hàng vào cơ sở dữ liệu
            database.them_kh(ten_kh, ngay_sinh, gioi_tinh, int(sdt), int(cccd), email, dia_chi)
            messagebox.showinfo("Thông báo", "Thêm khách hàng thành công")
            clear_entries()
            display_data()
            addThem.destroy()  # Ẩn cửa sổ sau khi thêm thành công
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi thêm khách hàng: {str(e)}")

    def xuatFile():
        try:
            # Lấy tất cả các hàng dữ liệu từ bảng
            rows = table_kh.get_children()
            if not rows:
                messagebox.showinfo("Thông báo", "Không có dữ liệu để xuất.")
                return

            # Tạo workbook mới và chọn worksheet hoạt động
            wb = Workbook()
            ws = wb.active
            ws.title = "KhachHang"

            # Thêm tiêu đề cột vào worksheet
            columns = ["ID", "Tên NV", "Ngày Sinh", "Giới Tính", "SĐT", "CCCD", "Email", "Địa Chỉ", "Ngày ĐK"]
            for col_num, column_title in enumerate(columns, 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}1'] = column_title

            # Thêm dữ liệu từ bảng vào worksheet
            for row_num, row_id in enumerate(rows, 2):
                row_data = table_kh.item(row_id, 'values')
                for col_num, cell_value in enumerate(row_data, 1):
                    column_letter = get_column_letter(col_num)
                    ws[f'{column_letter}{row_num}'] = cell_value

            # Lưu workbook vào file
            file_path = "DanhSachKhachHang.xlsx"
            wb.save(file_path)
            messagebox.showinfo("Thông báo", f"Xuất dữ liệu thành công vào file {file_path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi xuất dữ liệu: {str(e)}")
    def fromThem():
        def center_window(window, width, height):
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()

            x = (screen_width / 2) - (width / 2)
            y = (screen_height / 2) - (height / 2)
            window.geometry('%dx%d+%d+%d' % (width, height, x, y))

        addThem = Toplevel()
        addThem.title('Thêm Khách Hàng')
        addThem.geometry('500x600')
        center_window(addThem, 500, 600)
        addThem.transient(frame_noidung.winfo_toplevel())  # Đảm bảo cửa sổ luôn ở trên cùng của ứng dụng chính
        addThem.attributes("-topmost", True)  # Đảm bảo nó trên tất cả các cửa sổ khác
        addThem.focus_force()  # Đặt trọng tâm vào cửa sổ này

        image = Image.open("img/backgroud.jpg")
        global background_image
        background_image = ImageTk.PhotoImage(image)
        canvas = Canvas(addThem, width=addThem.winfo_screenwidth(), height=addThem.winfo_screenheight())
        canvas.pack(fill=BOTH, expand=YES)

        def resize_image(event):
            new_width = event.width
            new_height = event.height
            resized_image = image.resize((new_width, new_height), Image.LANCZOS)
            global background_image_resized
            background_image_resized = ImageTk.PhotoImage(resized_image)
            canvas.create_image(0, 0, anchor=NW, image=background_image_resized)
            canvas.image = background_image_resized

        canvas.bind("<Configure>", resize_image)
        frame = Frame(addThem, bg="white", width=400, height=500)
        frame.place(relx=0.5, rely=0.5, anchor=CENTER)

        lb = Label(frame, text="Thêm Khách Hàng", fg="gray", font=("Arial", 18, "bold"), bg='White')
        lb.place(x=88, y=10)

        lb_ht = Label(frame, text="Họ Tên", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_ht.place(x=30, y=70)

        lb_ns = Label(frame, text="Ngày Sinh", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_ns.place(x=30, y=120)

        lb_sex = Label(frame, text="Giới Tính", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_sex.place(x=30, y=170)

        lb_sdt = Label(frame, text="SĐT", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_sdt.place(x=30, y=220)

        lb_cccd = Label(frame, text="CCCD", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_cccd.place(x=30, y=270)

        lb_email = Label(frame, text="Email", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_email.place(x=30, y=320)

        lb_diachi = Label(frame, text="Địa chỉ", fg="gray", font=("Arial", 10, "bold"), bg='White')
        lb_diachi.place(x=30, y=370)

        # Nhap

        input_ten_kh = Entry(frame, font=("Arial", 12), highlightbackground="light blue", highlightthickness=1,
                             borderwidth=0)
        input_ten_kh.place(x=120, y=70)

        input_ngay_kh = DateEntry(frame, font=("Arial", 11), date_pattern="y-mm-dd", highlightbackground="light blue",
                                  highlightthickness=1,
                                  borderwidth=0, width=20)
        input_ngay_kh.place(x=120, y=120)

        input_sex_kh = ttk.Combobox(frame, state="readonly", font=("Arial", 11), width=18, values=('Nam', 'Nữ'))
        input_sex_kh.place(x=120, y=170)

        input_sdt_kh = Entry(frame, font=("Arial", 12), highlightbackground="light blue",
                             highlightthickness=1, borderwidth=0)
        input_sdt_kh.place(x=120, y=220)

        input_cccd_kh = Entry(frame, font=("Arial", 12), highlightbackground="light blue", highlightthickness=1,
                              borderwidth=0)
        input_cccd_kh.place(x=120, y=270)

        input_email_kh = Entry(frame, font=("Arial", 12), highlightbackground="light blue", highlightthickness=1,
                               borderwidth=0)
        input_email_kh.place(x=120, y=320)

        input_diachi_kh = Entry(frame, font=("Arial", 12), highlightbackground="light blue", highlightthickness=1,
                                borderwidth=0)
        input_diachi_kh.place(x=120, y=370)

        btn_them = Button(frame, text="Thêm",
                          command=lambda: them(input_ten_kh, input_ngay_kh, input_sex_kh, input_sdt_kh, input_cccd_kh,
                                               input_email_kh, input_diachi_kh, addThem), bg="#57a1f8", fg="White",
                          font=("Arial", 10, "bold"), padx=30, pady=3,
                          borderwidth=0)
        btn_them.place(x=50, y=430)

        btn_dangky = Button(frame, text="Hủy", bg="#57a1f8", fg="White", font=("Arial", 10, "bold"), padx=45, pady=3,
                            borderwidth=0, command=addThem.destroy)  # Đóng cửa sổ khi bấm Hủy
        btn_dangky.place(x=215, y=430)

        addThem.mainloop()

    frame_tb = Frame(frame_noidung, bg="White")
    frame_tb.pack(side="top", fill="y")

    frame_chucnang = Frame(frame_noidung, width=300, height=200, bg="White")
    frame_chucnang.pack(side="bottom", fill="both", expand=True)

    def update_suggestions(event):
        input_text = input_timkiem_kh.get().strip().lower()
        if input_text == "":
            display_data()  # Hiển thị lại toàn bộ dữ liệu nếu không có chuỗi tìm kiếm
            return

        rows = database.Hien_Thi_KH()
        filtered_rows = [row for row in rows if
                         (input_text in str(row[1]).lower()) or (str(row[5]).find(input_text) != -1)]

        update_table(filtered_rows)
    frame_timkiem = Frame(frame_tb, bg="White")
    frame_timkiem.pack(side="top", pady=10)

    input_timkiem_kh = ttk.Combobox(frame_timkiem, font=("Arial", 11)
                             , width=30)
    input_timkiem_kh.grid(row=0, column=0, padx=10, pady=10)
    input_timkiem_kh.bind('<KeyRelease>', update_suggestions)
    anh = Image.open("img/icontim.png")
    size = anh.resize((20, 20), Image.LANCZOS)
    global img
    img = ImageTk.PhotoImage(size)
    btn_timkiem_kh = Button(frame_timkiem, image=img, borderwidth=0)
    btn_timkiem_kh.grid(row=0, column=1, padx=(0, 10))

    columns = ("ma_kh", "Họ và tên", "Ngày Sinh", "Giới Tính", "Số điện thoại", "CCCD", "Email", "Địa Chỉ","Ngày ĐK")
    table_kh = ttk.Treeview(frame_tb, columns=columns, show="headings")
    table_kh.heading("ma_kh", text="ID")
    table_kh.heading("Họ và tên", text="Tên NV")
    table_kh.heading("Ngày Sinh", text="Ngày Sinh")
    table_kh.heading("Giới Tính", text="Giới Tính")
    table_kh.heading("Số điện thoại", text="SĐT")
    table_kh.heading("CCCD", text="CCCD")
    table_kh.heading("Email", text="Email")
    table_kh.heading("Địa Chỉ", text="Địa Chỉ")
    table_kh.heading("Ngày ĐK", text="Ngày ĐK")
    table_kh.column("ma_kh", minwidth=0, width=50)
    table_kh.column("Họ và tên", minwidth=0, width=150)
    table_kh.column("Ngày Sinh", minwidth=0, width=100)
    table_kh.column("Giới Tính", minwidth=0, width=80)
    table_kh.column("Số điện thoại", minwidth=0, width=120)
    table_kh.column("CCCD", minwidth=0, width=100)
    table_kh.column("Email", minwidth=0, width=150)
    table_kh.column("Địa Chỉ", minwidth=0, width=250)
    table_kh.column("Ngày ĐK", minwidth=0, width=100)
    table_kh.pack(fill="both", expand=True)
    for col in columns:
        table_kh.column(col, anchor='center')

    table_kh.bind("<<TreeviewSelect>>", ChonDuLieuBang)
    table_kh.pack(padx=10, pady=10, fill=BOTH, expand=True)
    ma_chon = None
    display_data()

    btn_them = Button(frame_chucnang, command=fromThem, text="Thêm", font=("Arial", 8, "bold"), fg="White", borderwidth=0,
                      bg="#57a1f8", pady=6, padx=45)
    btn_them.place(x=5, y=30)
    btn_sua = Button(frame_chucnang, command=sua, text="Sửa", font=("Arial", 8, "bold"), fg="White", borderwidth=0,
                     bg="#57a1f8", pady=6, padx=45)
    btn_sua.place(x=150, y=30)
    btn_xoa = Button(frame_chucnang, command=xoa, text="Xóa", font=("Arial", 8, "bold"), fg="White", borderwidth=0,
                     bg="#57a1f8", pady=6, padx=45)
    btn_xoa.place(x=290, y=30)
    btn_xuatbao = Button(frame_chucnang,command=xuatFile, text="Xuất Danh Sách", font=("Arial", 8, "bold"), fg="White", borderwidth=0,
                     bg="#57a1f8", pady=6, padx=45)
    btn_xuatbao.place(x=430, y=30)

    lb_ten_kh = Label(frame_chucnang, bg="White", text="Tên khách hàng", fg="black", font=("Arial", 10, "bold"))
    lb_ten_kh.place(x=5, y=70)
    input_ten_kh = Entry(frame_chucnang, font=("Arial", 11), highlightbackground="light blue", highlightthickness=1,
                         borderwidth=0, width=20)
    input_ten_kh.place(x=5, y=90)
    lb_ngay_kh = Label(frame_chucnang, bg="White", text="Ngày sinh", fg="black", font=("Arial", 10, "bold"))
    lb_ngay_kh.place(x=5, y=120)
    input_ngay_kh = DateEntry(frame_chucnang, font=("Arial", 11), date_pattern="y-mm-dd", highlightbackground="light blue", highlightthickness=1,
                          borderwidth=0, width=20)
    input_ngay_kh.place(x=5, y=140)

    lb_sex_kh = Label(frame_chucnang, bg="White", text="Giới tính", fg="black", font=("Arial", 10, "bold"))
    lb_sex_kh.place(x=200, y=70)
    input_sex_kh = ttk.Combobox(frame_chucnang, state="readonly", font=("Arial", 11), width=18, values=('Nam', 'Nữ'))
    input_sex_kh.place(x=200, y=90)

    lb_sdt_kh = Label(frame_chucnang, bg="White", text="Số điện thoại", fg="black", font=("Arial", 10, "bold"))
    lb_sdt_kh.place(x=200, y=120)
    input_sdt_kh = Entry(frame_chucnang, bg="White", font=("Arial", 11), highlightbackground="light blue", highlightthickness=1,
                         borderwidth=0, width=20)
    input_sdt_kh.place(x=200, y=140)

    lb_cccd_kh = Label(frame_chucnang, bg="White", text="Chứng minh thư", fg="black", font=("Arial", 10, "bold"))
    lb_cccd_kh.place(x=395, y=70)
    input_cccd_kh = Entry(frame_chucnang, bg="White", font=("Arial", 11), highlightbackground="light blue", highlightthickness=1,
                          borderwidth=0, width=20)
    input_cccd_kh.place(x=395, y=90)

    lb_email_kh = Label(frame_chucnang, bg="White", text="Email", fg="black", font=("Arial", 10, "bold"))
    lb_email_kh.place(x=395, y=120)
    input_email_kh = Entry(frame_chucnang, font=("Arial", 11), highlightbackground="light blue", highlightthickness=1,
                           borderwidth=0, width=20)
    input_email_kh.place(x=395, y=140)

    lb_diachi_kh = Label(frame_chucnang, bg="White", text="Địa chỉ", fg="black", font=("Arial", 10, "bold"))
    lb_diachi_kh.place(x=590, y=70)
    input_diachi_kh = Entry(frame_chucnang, font=("Arial", 11), highlightbackground="light blue", highlightthickness=1,
                            borderwidth=0, width=20)
    input_diachi_kh.place(x=590, y=90)
